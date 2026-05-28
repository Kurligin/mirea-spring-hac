import io
from datetime import datetime
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_admin, get_db
from app.core.permissions import assert_can_access_event
from app.models.admin_account import AdminAccount
from app.models.event import Event
from app.models.event_field import EventField
from app.models.registration import Registration, RegistrationStatus
from app.models.user import User
from app.schemas.registration import RegistrationResponse

router = APIRouter(prefix="/api/admin/events", tags=["admin-registrations"])


@router.get("/{event_id}/registrations", response_model=list[RegistrationResponse])
async def list_registrations(
    event_id: UUID,
    status: RegistrationStatus | None = Query(None),
    q: str | None = Query(
        None,
        description="Поиск по short_code, ФИО или username",
    ),
    limit: int = Query(500, le=2000),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    admin: AdminAccount = Depends(get_current_admin),
):
    event = (await db.execute(select(Event).where(Event.id == event_id))).scalar_one_or_none()
    if event is None:
        raise HTTPException(404, "Event not found")
    assert_can_access_event(admin, event)

    stmt = (
        select(Registration, User)
        .join(User, User.id == Registration.user_id)
        .where(Registration.event_id == event_id)
        .order_by(Registration.created_at.desc())
        .limit(limit)
        .offset(offset)
    )
    if status is not None:
        stmt = stmt.where(Registration.status == status)
    if q:
        needle = q.strip()
        like = f"%{needle.lower()}%"
        stmt = stmt.where(
            or_(
                Registration.short_code == needle.upper(),
                User.first_name.ilike(like),
                User.last_name.ilike(like),
                User.username.ilike(like),
            )
        )
    rows = list((await db.execute(stmt)).all())

    out: list[RegistrationResponse] = []
    for reg, user in rows:
        resp = RegistrationResponse.model_validate(reg)
        resp.user_first_name = user.first_name
        resp.user_last_name = user.last_name
        resp.user_username = user.username
        out.append(resp)
    return out


@router.get("/{event_id}/registrations.xlsx")
async def export_registrations_xlsx(
    event_id: UUID,
    db: AsyncSession = Depends(get_db),
    admin: AdminAccount = Depends(get_current_admin),
):
    """XLSX-экспорт всех регистраций события (с ответами анкеты)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    event = (await db.execute(select(Event).where(Event.id == event_id))).scalar_one_or_none()
    if event is None:
        raise HTTPException(404, "Event not found")
    assert_can_access_event(admin, event)

    fields = list(
        (await db.execute(
            select(EventField).where(EventField.event_id == event_id).order_by(EventField.position)
        )).scalars().all()
    )
    rows = list(
        (await db.execute(
            select(Registration, User)
            .join(User, User.id == Registration.user_id)
            .where(Registration.event_id == event_id)
            .order_by(Registration.created_at.asc())
        )).all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Регистрации"
    headers = [
        "Код", "Статус", "Имя", "Фамилия", "Username",
        "Создано", "Чек-ин", "Поздняя отмена",
    ] + [f.label for f in fields]
    ws.append(headers)
    header_fill = PatternFill(start_color="2C54EE", end_color="2C54EE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def _fmt_dt(dt: datetime | None) -> str:
        return dt.astimezone().strftime("%d.%m.%Y %H:%M") if dt else ""

    for reg, user in rows:
        row = [
            reg.short_code or "",
            reg.status.value,
            user.first_name or "",
            user.last_name or "",
            user.username or "",
            _fmt_dt(reg.created_at),
            _fmt_dt(reg.checked_in_at),
            "да" if reg.is_late_cancellation else "",
        ]
        for f in fields:
            v = (reg.answers or {}).get(f.key)
            if isinstance(v, list):
                v = ", ".join(str(x) for x in v)
            elif isinstance(v, bool):
                v = "да" if v else "нет"
            row.append("" if v is None else str(v))
        ws.append(row)

    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_title = "".join(c for c in event.title if c.isalnum() or c in "-_ ")[:60].strip() or "event"
    filename = f"{safe_title}-{event_id.hex[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
